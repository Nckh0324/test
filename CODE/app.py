import sys
import cv2
import face_recognition
import os
import numpy as np
from datetime import datetime
import openpyxl
#GUI
import customtkinter as ctk
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
from tkcalendar import DateEntry
from PIL import Image, ImageTk
#Sms
from twilio.rest import Client

import pandas as pd
#Gmail
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText

import logging

if getattr(sys, 'frozen', False):
    # Đường dẫn tới thư mục chứa executable
    application_path = sys._MEIPASS
else:
    application_path = os.path.dirname(os.path.abspath(__file__))

class CameraApp:
    def __init__(self):
        self.window = ctk.CTk()
        self.window.title("HỆ THỐNG ĐIỂM DANH THÔNG QUA CAMERA")
        self.window.geometry("1500x850")
        self.window.resizable(False, False)

        self.video_capture = None
        self.images = []  # Thêm dòng này để khởi tạo self.images
        self.encoded_known_faces = []
        self.known_names = []
        self.dataframe = None
        self.last_attendance_time = {}
        self.attendance_cooldown = 5
        self.excel_data = None
        self.selected_date = None
        self.excel_file = None # Khởi tạo biến excel_file
        
        # Đường dẫn thư mục ảnh mặc định
        self.default_faces_directory = r"D:\python\DATA" # Thay đổi đường dẫn nếu cần
        
        # Tự động tải ảnh khuôn mặt khi khởi động
        self.load_images(self.default_faces_directory)
        self.encode_faces()
        
        # Set custom appearance
        ctk.set_appearance_mode("light")
        ctk.set_default_color_theme("blue")
        
        # Create GUI
        self.create_gui()

            
    def create_gui(self):
        # Title Label with large font
        title_label = ctk.CTkLabel(
            self.window,
            text="HỆ THỐNG ĐIỂM DANH THÔNG QUA CAMERA",
            font=("Helvetica", 24, "bold")
        )
        title_label.pack(pady=20)

        # Main content frame
        content_frame = ctk.CTkFrame(self.window)
        content_frame.pack(expand=True, fill="both", padx=20, pady=20)

        # Left section - Camera
        camera_section = ctk.CTkFrame(content_frame)
        camera_section.pack(side="left", padx=10, pady=10, fill="both")

        camera_label = ctk.CTkLabel(
            camera_section,
            text="Hệ thống CAMERA",
            font=("Helvetica", 16, "bold")
        )
        camera_label.pack(pady=10)

        # Camera Canvas
        self.label_camera = ctk.CTkLabel(camera_section, text="", width=640, height=480)
        self.label_camera.pack(padx=10, pady=10)
        

        # Camera controls
        camera_controls = ctk.CTkFrame(camera_section)
        camera_controls.pack(fill="x", padx=10, pady=10)

        id_label = ctk.CTkLabel(camera_controls, text="ID")
        id_label.pack(side="left", padx=5)

        self.camera_id = ctk.CTkEntry(camera_controls, width=100)
        self.camera_id.pack(side="left", padx=5)

        self.btn_start_stop = ctk.CTkButton(
            camera_controls,
            text="Bật Camera",
            command=self.toggle_camera,
            width=120
        )
        self.btn_start_stop.pack(side="left", padx=5)

        # Right section
        right_section = ctk.CTkFrame(content_frame)
        right_section.pack(side="left", padx=10, pady=10, fill="both", expand=True)

        # Excel file selection
        excel_frame = ctk.CTkFrame(right_section)
        excel_frame.pack(fill="x", pady=10)

        excel_label = ctk.CTkLabel(
            excel_frame,
            text="Chọn File excel để điểm danh",
            font=("Helvetica", 14)
        )
        excel_label.pack(pady=5)

        excel_controls = ctk.CTkFrame(excel_frame)
        excel_controls.pack(fill="x")

        self.excel_entry = ctk.CTkEntry(
            excel_controls,
            placeholder_text="Chưa chọn file Excel",
            width=300
        )
        self.excel_entry.pack(side="left", padx=5)

        excel_button = ctk.CTkButton(
            excel_controls,
            text="Chọn file Excel",
            command=self.select_excel_file,
            width=120
        )
        excel_button.pack(side="left", padx=5)

        display_excel_button = ctk.CTkButton(
            excel_controls,
            text="Hiển thị file Excel",
            command=self.display_excel_data,
            width=120
        )
        display_excel_button.pack(side="left", padx=5)
        
        # button send email to parents
        send_email_button = ctk.CTkButton(
            excel_controls,
            text="Gửi email cho phụ huynh",
            command=self.send_email_to_parents,
            width=180
        )
        send_email_button.pack(side="left", padx=5)

        # Date selection
        date_frame = ctk.CTkFrame(right_section)
        date_frame.pack(fill="x", pady=10)

        date_label = ctk.CTkLabel(
            date_frame,
            text="Ngày điểm danh",
            font=("Helvetica", 14)
        )
        date_label.pack(pady=5)

        date_controls = ctk.CTkFrame(date_frame)
        date_controls.pack(fill="x")

        self.date_entry = DateEntry(
            date_controls,
            width=30,
            background='darkblue',
            foreground='white',
            borderwidth=2,
            date_pattern='dd/mm/yyyy'
        )
        self.date_entry.pack(side="left", padx=5)

        date_button = ctk.CTkButton(
            date_controls,
            text="Chọn ngày",
            command=self.get_selected_date,
            width=120
            )
        date_button.pack(side="left", padx=5)

        # Create Treeview frame
        treeview_frame = ctk.CTkFrame(right_section)
        treeview_frame.pack(fill="both", expand=True, pady=10)

        # Create Treeview
        self.tree = ttk.Treeview(treeview_frame, selectmode='browse')
        self.tree.pack(fill="both", expand=True, padx=5, pady=5)

        # Add scrollbar to treeview
        scrollbar = ttk.Scrollbar(treeview_frame, orient="vertical", command=self.tree.yview)
        scrollbar.pack(side="right", fill="y")
        self.tree.configure(yscrollcommand=scrollbar.set)

        # Horizontal scrollbar
        h_scrollbar = ttk.Scrollbar(treeview_frame, orient="horizontal", command=self.tree.xview)
        h_scrollbar.pack(side="bottom", fill="x")
        self.tree.configure(xscrollcommand=h_scrollbar.set)

        # Style configuration for Treeview
        style = ttk.Style()
        style.configure("Treeview",
                       background="white",
                       foreground="black",
                       rowheight=25,
                       fieldbackground="white")
        style.configure("Treeview.Heading",
                       background="lightgray",
                       foreground="black",
                       relief="raised")

        # Buttons frame
        buttons_frame = ctk.CTkFrame(right_section)
        buttons_frame.pack(fill="x", pady=10, padx=10)

        unmarked_button = ctk.CTkButton(
            buttons_frame,
            text="Danh sách học sinh\nchưa điểm danh",
            command=self.show_unmarked,
            height=60
        )
        unmarked_button.pack(side="left", padx=5, expand=True)

        marked_button = ctk.CTkButton(
            buttons_frame,
            text="Danh sách học sinh\nđã điểm danh",
            command=self.show_marked,
            height=60
        )
        marked_button.pack(side="left", padx=5, expand=True)

        clear_button = ctk.CTkButton(
            buttons_frame,
            text="Xóa nội dung",
            command=self.clear_log,
            height=60
        )
        clear_button.pack(side="left", padx=5, expand=True)

    def get_selected_date(self):
        self.selected_date = self.date_entry.get_date().strftime("%d/%m/%Y")
        print(f"Đã chọn ngày: {self.selected_date}")

    def load_images(self, image_directory):
        # Reset danh sách ảnh
        self.images = []
        self.known_names = []
        
        # Kiểm tra thư mục tồn tại
        if not os.path.exists(image_directory):
            messagebox.showerror("Lỗi", f"Thư mục {image_directory} không tồn tại")
            return
        
        # Đọc danh sách ảnh
        image_list = os.listdir(image_directory)
        for image_file in image_list:
            current_image = cv2.imread(os.path.join(image_directory, image_file))
            if current_image is not None:
                self.images.append(current_image)
                self.known_names.append(os.path.splitext(image_file)[0])
    
    def encode_faces(self):
        # Reset danh sách mã hóa khuôn mặt
        self.encoded_known_faces = []
        
        for img in self.images:
            rgb_image = cv2.cvtColor(img, cv2.COLOR_BGR2RGB)
            encodings = face_recognition.face_encodings(rgb_image)
            if encodings:
                self.encoded_known_faces.append(encodings[0])
    
    def toggle_camera(self):
        if not self.excel_file:
            messagebox.showwarning("Lỗi", "Chưa chọn file Excel.")
            return 
        
        camera_index = self.camera_id.get().strip()
        
        try: 
            camera_index = int(camera_index)
        except ValueError: 
            messagebox.showwarning("Lỗi", "Giá trị ID không hợp lệ")
            return

        if self.video_capture is None: 
            self.video_capture = cv2.VideoCapture(camera_index) 
            if not self.video_capture.isOpened():
                messagebox.showerror("Lỗi", "Không thể mở camera")
                return
            
            self.btn_start_stop.configure(text="Tắt Camera")
            self.update_frame()
        else: 
            if hasattr(self, 'after_id'):
                self.window.after_cancel(self.after_id)
            
            self.video_capture.release()
            self.label_camera.configure(image='')
            self.video_capture = None 
            self.btn_start_stop.configure(text="Bật Camera")
    
    def update_frame(self):
        ret, frame = self.video_capture.read()
        if not ret:
            messagebox.showerror("Lỗi", "Không thể đọc khung hình từ webcam")
            return

        small_frame = cv2.resize(frame, (0, 0), None, fx=0.5, fy=0.5)
        rgb_small_frame = cv2.cvtColor(small_frame, cv2.COLOR_BGR2RGB)

        current_face_locations = face_recognition.face_locations(rgb_small_frame)
        current_face_encodings = face_recognition.face_encodings(rgb_small_frame, current_face_locations)

        for face_encoding, face_location in zip(current_face_encodings, current_face_locations):
            matches = face_recognition.compare_faces(self.encoded_known_faces, face_encoding)
            face_distances = face_recognition.face_distance(self.encoded_known_faces, face_encoding)
            best_match_index = np.argmin(face_distances)

            if face_distances[best_match_index] < 0.50:
                name = self.known_names[best_match_index].lower()
                self.mark_attendance([name]) # Chú ý: truyền list [name] thay vì name
            else:
                name = "Unknown"

            top, right, bottom, left = face_location
            top, right, bottom, left = top * 2, right * 2, bottom * 2, left * 2
            cv2.rectangle(frame, (left, top), (right, bottom), (0, 255, 0), 2)
            cv2.putText(frame, name, (left + 6, bottom - 6), cv2.FONT_HERSHEY_COMPLEX, 1, (255, 255, 255), 2)

        # Chuyển đổi frame sang định dạng ảnh cho Tkinter
        frame = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
        img = Image.fromarray(frame)
        imgtk = ImageTk.PhotoImage(image=img)
        self.label_camera.configure(image=imgtk)
        self.label_camera.image = imgtk  # Giữ tham chiếu
        
        # Lên lịch update frame tiếp theo
        self.after_id = self.window.after(20, self.update_frame)

    def mark_attendance(self, student_IDs):
        # Kiểm tra điều kiện ban đầu
        if not hasattr(self, 'excel_file') or not self.excel_file:
            messagebox.showwarning("Lỗi", "Chưa chọn file Excel")
            return
        
        # Đảm bảo có ngày được chọn
        if not hasattr(self, 'selected_date') or not self.selected_date:
            self.selected_date = self.date_entry.get_date().strftime("%d/%m/%Y")
        
        try:
            workbook = openpyxl.load_workbook(self.excel_file)
            sheet = workbook.active

            column_letter = None
            for col in range(1, sheet.max_column + 1):
                cell_value = sheet.cell(row=1, column=col).value
                if cell_value == self.selected_date:
                    column_letter = col
                    break

            # Nếu chưa có cột cho ngày này, tạo cột mới
            if column_letter is None:
                column_letter = sheet.max_column + 1
                sheet.cell(row=1, column=column_letter, value=self.selected_date)

            # Lấy thời gian điểm danh hiện tại
            now = datetime.now()
            datetime_string = now.strftime("%H:%M:%S")

            # Biến lưu kết quả
            newly_marked_ids = []
            missing_ids = []
            already_marked_ids = []

            # Đảm bảo student_IDs là list
            if not isinstance(student_IDs, list):
                student_IDs = [student_IDs]

            # Chuyển đổi sang string và loại bỏ khoảng trắng
            student_IDs = [str(sid).strip() for sid in student_IDs]
            
            # Debug log
            print(f"Attempting to mark attendance for: {student_IDs}")

            # Duyệt qua từng mã sinh viên
            for student_ID_str in student_IDs:
                updated = False
                
                # Duyệt qua các hàng để tìm sinh viên
                for row in range(2, sheet.max_row + 1):  # Bắt đầu từ hàng 2
                    cell_name = sheet.cell(row=row, column=1).value
                    
                    # So sánh mã sinh viên
                    if cell_name and str(cell_name).strip().lower() == student_ID_str.strip().lower():
                        # Kiểm tra ô điểm danh
                        attendance_cell = sheet.cell(row=row, column=column_letter)
                        
                        if attendance_cell.value:  # Đã điểm danh
                            already_marked_ids.append(student_ID_str)
                            break
                        else:
                            # Điểm danh
                            attendance_cell.value = f"Đã điểm danh - {datetime_string}"
                            newly_marked_ids.append(student_ID_str)
                            updated = True
                            break
                
                # Nếu không tìm thấy sinh viên
                if not updated:
                    missing_ids.append(student_ID_str)

            # Tạo báo cáo kết quả
            report = []
            if newly_marked_ids:
                report.append(f"Điểm danh thành công:\n{', '.join(newly_marked_ids)}")
                # Thông báo popup
                messagebox.showinfo("Điểm Danh", "\n".join(newly_marked_ids))
            if already_marked_ids:
                report.append(f"Đã điểm danh trước đó:\n{', '.join(already_marked_ids)}")
            if missing_ids:
                report.append(f"Không tìm thấy trong danh sách lớp:\n{', '.join(missing_ids)}")

            # In báo cáo ra console
            if report:
                print("\n".join(report))

            # Lưu file Excel
            workbook.save(self.excel_file)
            workbook.close()
            print("Cập nhật file Excel thành công.")
            
        except Exception as e:
            messagebox.showerror("Lỗi", f"Lỗi trong quá trình xử lý file Excel: {str(e)}")
            print(f"Lỗi chi tiết: {e}")
            import traceback
            traceback.print_exc()

    def select_excel_file(self):
        filepath = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
        if filepath:
            self.excel_file = filepath
            self.excel_entry.delete(0, "end")
            self.excel_entry.insert(0, os.path.basename(filepath))         
            
    def display_excel_data(self):
        if not self.excel_file:
            messagebox.showwarning("Lỗi", "Vui lòng chọn file Excel trước.")
            return

        try:
            self.dataframe = pd.read_excel(self.excel_file)

            # Clear existing items
            for item in self.tree.get_children():
                self.tree.delete(item)

            # Update columns for Treeview
            self.tree['columns'] = tuple(self.dataframe.columns)
            self.tree.column("#0", width=0, stretch=tk.NO)  # Hide the first column
            for column in self.dataframe.columns:
                self.tree.column(column, anchor=tk.W, width=150)
                self.tree.heading(column, text=column, anchor=tk.W)

            # Add rows
            for _, row in self.dataframe.iterrows():
                values = [row[col] for col in self.dataframe.columns]
                self.tree.insert("", tk.END, values=values)
        except Exception as e:
            messagebox.showerror("Lỗi", f"Lỗi khi tải file Excel: {str(e)}")

    def show_unmarked(self):
        if self.dataframe is None:
            messagebox.showwarning("Lỗi", "Chưa tải dữ liệu từ file Excel.")
            return

        try:
            self.selected_date = self.date_entry.get_date().strftime("%d/%m/%Y")
            
            if self.selected_date not in self.dataframe.columns:
                messagebox.showerror("Lỗi", "Ngày được chọn chưa tồn tại trong dữ liệu.")
                return

            unmarked_students = self.dataframe[pd.isna(self.dataframe[self.selected_date])]
            self.update_treeview(unmarked_students)
        
        except Exception as e:
            messagebox.showerror("Lỗi", str(e))

    def show_marked(self):
        if self.dataframe is None:
            messagebox.showwarning("Lỗi", "Chưa tải dữ liệu từ file Excel.")
            return

        try:
            self.selected_date = self.date_entry.get_date().strftime("%d/%m/%Y")
            
            if self.selected_date not in self.dataframe.columns:
                messagebox.showerror("Lỗi", "Ngày được chọn chưa tồn tại trong dữ liệu.")
                return

            marked_students = self.dataframe[pd.notna(self.dataframe[self.selected_date])]
            self.update_treeview(marked_students)
        
        except Exception as e:
            messagebox.showerror("Lỗi", str(e))

    def send_email_to_parents(self):
        if self.dataframe is None:
            messagebox.showwarning("Lỗi", "Chưa tải dữ liệu từ file Excel.")
            return

        try:
            selected_date = self.date_entry.get_date().strftime("%d/%m/%Y")
            
            # Kiểm tra cột email
            if len(self.dataframe.columns) < 5:
                messagebox.showerror("Lỗi", "File Excel không đủ cột.")
                return

            email_column = self.dataframe.columns[4]  # Cột email
            name_column = self.dataframe.columns[1]   # Cột tên

            # Lọc các học sinh vắng mặt
            unmarked_students = self.dataframe[pd.isna(self.dataframe[selected_date])]

            if unmarked_students.empty:
                messagebox.showinfo("Thông báo", "Không có học sinh vắng mặt.")
                return

            # Cấu hình email
            sender_email = "vobaolong15@gmail.com"
            sender_password = "wpmrfsgbhsfgwiin"

            # Gửi email cho từng phụ huynh
            for _, student in unmarked_students.iterrows():
                parent_email = student[email_column]
                student_name = student[name_column]

                if pd.notna(parent_email):
                    try:
                        # Tạo email
                        msg = MIMEMultipart()
                        msg['From'] = sender_email
                        msg['To'] = parent_email
                        msg['Subject'] = f"Thông báo vắng học của {student_name}"

                        body = f"""
                        Kính gửi Quý Phụ huynh,

                        Nhà trường xin thông báo: Em {student_name} đã vắng mặt vào ngày {selected_date}.

                        Trân trọng,
                Trường THPT Nguyễn Thượng Hiền
                        """

                        msg.attach(MIMEText(body, 'plain', 'utf-8'))

                        # Kết nối và gửi email
                        with smtplib.SMTP('smtp.gmail.com', 587) as server:
                            server.starttls()
                            server.login(sender_email, sender_password)
                            server.send_message(msg)

                        print(f"Đã gửi email tới {parent_email}")

                    except Exception as e:
                        print(f"Lỗi gửi email tới {parent_email}: {e}")

            messagebox.showinfo("Thành công", "Đã gửi email thông báo.")

        except Exception as e:
            messagebox.showerror("Lỗi", str(e))

    def update_treeview(self, filtered_dataframe):
        # Xóa các item hiện tại
        for item in self.tree.get_children():
            self.tree.delete(item)

        # Thêm dữ liệu đã lọc
        for _, row in filtered_dataframe.iterrows():
            values = [str(row[col]) for col in self.dataframe.columns]
            self.tree.insert("", tk.END, values=values)

    def clear_log(self):
        for item in self.tree.get_children():
            self.tree.delete(item)
            
def main():
    app = CameraApp()
    app.window.mainloop()

if __name__ == "__main__":
    main()